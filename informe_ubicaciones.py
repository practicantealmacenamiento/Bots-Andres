from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font, NamedStyle, Font, Alignment, Side, Border, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from sap_gui_library import SapGui,Transaction,DataProcess
from datetime import date, datetime
from Generales import send_email, close_excel, report_status, create_folder, automatic_width
import matplotlib.pyplot as plt
import pandas as pd
import config, sqlite3, os, math, locale


#AFK CONSTANTS
USER = config.SAP_USERNAME 
PASSWORD = config.SAP_PASSWORD

#INITIAL CONFIG
SEND_EMAILS = True
AFK = True
BUSY_LOCATIONS = False
INSERT= True

#CONSTANTS
FOLDER_PATH = r"C:\TEMP\Informe de ubicaciones"
EXCEL_FILE = FOLDER_PATH + r"\Informe de ubicaciones.xlsx" 
SOURCE_FILE = FOLDER_PATH + r"\No facturar.xlsx"
DATABASE_PATH = r"C:\Users\prac.almacenamiento\OneDrive - Prebel S.A\Escritorio\Migue\Bots\Databases\Informe_ubicaciones.db"




# Script para crear la base de datos
def createDatabase():
    connection = sqlite3.connect(DATABASE_PATH)
    cursor = connection.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ocupacion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            ubicaciones_ocupadas INTEGER NOT NULL
        )'''
    )

    connection.commit()
    connection.close()


# Obtiene la informacion de la base de datos 
def fetchAll():
    connection = sqlite3.connect(DATABASE_PATH)
    cursor = connection.cursor()
    
    # Obtiene el mes actual para realizar la query
    current_month = datetime.now().strftime("%Y-%m")
    
    # Trae todos los valores del mes actual
    cursor.execute('''
        SELECT * 
        FROM ocupacion
        WHERE strftime('%Y-%m', fecha) = ?
    ''', (current_month,))
    
    records = cursor.fetchall()
    connection.close()
    
    return records


# Inserta valores a la base de datos
def insert( date, busy_locations):
    try:
        connection = sqlite3.connect(DATABASE_PATH)
        cursor = connection.cursor()
        
        cursor.execute('''
            INSERT INTO ocupacion (fecha, ubicaciones_ocupadas) 
            VALUES (?, ?)
            ''', 
            (date, busy_locations)
        )
        
        connection.commit()

    except sqlite3.Error as e:
        print(f"Error al insertar datos: {e}")

    finally:
        if connection:
            connection.close()


# Ejecuta la LX02 (DATOS PRINCIPALES)
def LX02( sap_instance: SapGui ) -> pd.DataFrame:
    transaction = Transaction(
        session = sap_instance.get_session(),
        code = "LX02"
    )
    
    transaction.start_transaction()

    session = sap_instance.get_session()
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/usr/ctxtS1_LGNUM").text = "exp"
    session.findById("wnd[0]/usr/ctxtS1_LGNUM").caretPosition = 3
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtS1_LGTYP-LOW").text = "rpt"
    session.findById("wnd[0]/usr/ctxtS1_LGTYP-LOW").setFocus()
    session.findById("wnd[0]/usr/ctxtS1_LGTYP-LOW").caretPosition = 3
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtWERKS-LOW").text = "1000"
    session.findById("wnd[0]/usr/ctxtWERKS-LOW").setFocus()
    session.findById("wnd[0]/usr/ctxtWERKS-LOW").caretPosition = 4
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtS1_LGPLA-LOW").setFocus()
    session.findById("wnd[0]/usr/ctxtS1_LGPLA-LOW").caretPosition = 0
    session.findById("wnd[0]/usr/btn%_S1_LGPLA_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpNOSV").select()
    session.findById("wnd[1]/tbar[0]/btn[24]").press()
    session.findById("wnd[1]/tbar[0]/btn[8]").press()

    

    transaction.run_transaction('Execute')
    
    transaction.export_in_toolbar(9)
    
    transaction.select_export_and_download(
        address = FOLDER_PATH,
        name = "informe de ubicaciones",
        type_export = "1",
        selection = 'Replace'
    )
    
    data = DataProcess(
        address_file = os.path.join(FOLDER_PATH, "informe de ubicaciones.csv"),
        skiprows = 5
    )
    
    data.edit_file(
        skip_first_row = True,
        strip_columns = True,
        dropna_columns = ['Material'],
        columns_from_str_to_float = ['St. disp.']
    )
    
    return data.get_df() 


# Ejecuta la ZPP56 (COLUMNA CODIGO)
def ZPP56( sap_instance: SapGui, main_df: pd.DataFrame ) -> pd.DataFrame:
    main_df["Material"] = main_df["Material"].str.slice(1) 
    
    transaction = Transaction(
        session=sap_instance.get_session(),
        code= "ZPP56"
    )

    transaction.start_transaction()

    transaction.session.findById("wnd[0]/tbar[0]/okcd").text = "/NZPP56"
    transaction.session.findById("wnd[0]").sendVKey(0)
    transaction.session.findById("wnd[0]/usr/ctxtSP$00008-LOW").text = "1000"
    transaction.session.findById("wnd[0]").sendVKey(0)
    transaction.session.findById("wnd[0]/usr/ctxtSP$00020-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00002-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00025-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00003-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00006-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00005-LOW").setFocus()
    transaction.session.findById("wnd[0]/usr/ctxtSP$00005-LOW").caretPosition = 8
    transaction.session.findById("wnd[0]/usr/btn%_SP$00005_%_APP_%-VALU_PUSH").press()

    transaction.set_values_in_multiple_selection_file(
        address = FOLDER_PATH,
        df = main_df,
        column = "Material",
        name = "df_ZPP56",
        clean_before = True
    )

    transaction.session.findById("wnd[0]/usr/txtSP$00051-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/txtSP$00055-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00034-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00009-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00013-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00010-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00011-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00012-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00029-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00007-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00054-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxtSP$00015-LOW").text = ""
    transaction.session.findById("wnd[0]/usr/ctxt%LAYOUT").text = "/INF_INV_ANT"
    transaction.session.findById("wnd[0]/usr/ctxt%LAYOUT").setFocus()
    transaction.session.findById("wnd[0]/usr/ctxt%LAYOUT").caretPosition = 12
    
    transaction.run_transaction("Execute")

    transaction.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarContextButton("&MB_VIEW")
    transaction.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem("&PRINT_BACK_PREVIEW")

    transaction.export_in_toolbar(45)

    transaction.select_export_and_download(
        address = FOLDER_PATH,
        name = "ZPP56"
    )

    data = DataProcess(
        address_file = os.path.join(FOLDER_PATH, "ZPP56.csv"),
        skiprows = 4
    )

    data.edit_file(
        skip_first_row = True,
    )

    return data.get_df()


# Agrega el campo codigo al DF usando como clave foranea la referencia
def processData( main_df: pd.DataFrame, df_zpp56: pd.DataFrame):
    data = pd.merge( main_df, df_zpp56, how = "left", on="Material" )
    
    return data

#Pasa la informacion a excel
def mainDataToExcel( df: pd.DataFrame, date:date ):
    df["Fecha"] = date.strftime("%d/%m/%Y")
    
    df["Material"] = "M" + df["Material"]
    
    df = df.reindex( columns = ["Fecha", "Material", "NºMaterial antiguo", "Lote", "Texto breve de material", "Tp.", "Ubicación", "St. disp.", "UMB", "Cad./FPC", "Fecha EM"])
    df.rename(
        inplace = True,
        columns = {
            'St. disp.':'cantidad',
            'Material':'Referencia',
            'NºMaterial antiguo': 'Codigo',
        }
    )
    
    df["Cad./FPC"] = df["Cad./FPC"].str.replace('.', '/', regex=False)
    df["Fecha EM"] = df["Fecha EM"].str.replace('.', '/', regex=False)
    
    df['Cad./FPC'] = pd.to_datetime(df['Cad./FPC'], format='%d/%m/%Y')
    df['Fecha EM'] = pd.to_datetime(df['Fecha EM'], format='%d/%m/%Y')
    
    
    df.to_excel(os.path.join(FOLDER_PATH, "Informe de ubicaciones.xlsx"), index = False)


# genera el resumen de ubicaciones usadas
def SummaryToExcel( df: pd.DataFrame, date: date):
    df = df.drop_duplicates(["Ubicación"])
    
    busy_locations = len(df)
    
    alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    workbook = load_workbook(EXCEL_FILE)
    worksheet = workbook['Sheet1']
    
    worksheet['O1'].value = "Total ubicaciones"
    worksheet['O1'].border = border
    worksheet['O1'].alignment = alignment
    worksheet['O1'].font = Font(bold = True)
    worksheet['P1'].value = busy_locations
    worksheet['P1'].alignment = alignment
    worksheet['P1'].border = border
    worksheet['P1'].font = Font(bold = True)
    
    if BUSY_LOCATIONS:
        worksheet.merge_cells('O5:P5')
        
        worksheet['O5'].value = "Ubicaciones ocupadas"
        worksheet['O5'].border = border
        worksheet['P5'].border = border
        worksheet['O5'].alignment = alignment
        worksheet['P5'].alignment = alignment
        
        index = 6
        for aux, row in df.iterrows():
            if worksheet[f'O{index}'].value == None:
                worksheet[f'O{index}'].value = row["Ubicación"]
                worksheet[f'O{index}'].border = border
                worksheet[f'O{index}'].alignment = alignment
                
            else:
                worksheet[f'P{index}'].value = row["Ubicación"]
                worksheet[f'P{index}'].border = border
                worksheet[f'P{index}'].alignment = alignment
                index+=1
    
    workbook.save(EXCEL_FILE)
    
    if INSERT:
        insert(
            date = date,
            busy_locations = busy_locations
        )




def add_no_facturar_sheet():
    # Cargar el archivo principal
    workbook = load_workbook(EXCEL_FILE)

    # Cargar el archivo "No facturar.xlsx" como un DataFrame
    if os.path.exists(SOURCE_FILE):
        df_no_facturar = pd.read_excel(SOURCE_FILE)
        
        # Crear una nueva hoja en el workbook
        sheet_name = "No Facturar"
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]  # Eliminar la hoja si ya existe para actualizarla
        worksheet = workbook.create_sheet(title=sheet_name)

        # Definir estilos
        bold_font = Font(bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        green = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        
        # Se cambia formato de fecha 
        date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
        if "date_style" not in workbook.named_styles:
            workbook.add_named_style(date_style)
            
        # Escribir encabezados con estilos
        for col_idx, column_name in enumerate(df_no_facturar.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=column_name)
            cell.font = bold_font
            cell.alignment = alignment
            cell.border = border
            cell.fill = green  # Fondo verde para encabezados
            

        # Escribir el contenido del DataFrame
        for r_idx, row in enumerate(df_no_facturar.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border  # Aplicar bordes a todas las celdas

        # Ajustar el ancho de las columnas automáticamente
        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Congelar la primera fila
        worksheet.freeze_panes = 'A2'

        # Guardar los cambios en el archivo principal
        workbook.save(EXCEL_FILE)
        print("Hoja 'No Facturar' añadida y estilizada correctamente.")
    else:
        print(f"El archivo {SOURCE_FILE} no existe.")



# Estiliza la hoja de excel
def addStyles():
    bold_font = Font(bold=True)

    alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    
    red = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type ="solid")
    green = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
    blue = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    workbook= load_workbook(EXCEL_FILE)
    worksheet = workbook['Sheet1']
    
    # ----------------------- TEMPORAL -----------------------------------
    # worksheet['L1'].value = "Factura"
    # worksheet['L1'].border = border
    # worksheet['L1'].alignment = alignment
    # worksheet['L1'].font = bold_font
    
    # worksheet['M1'].value = "Orden de compra"
    # worksheet['M1'].border = border
    # worksheet['M1'].alignment = alignment
    # worksheet['M1'].font = bold_font
    #---------------------------------------------------------------------
    
    worksheet.title = "Informe de ubicaciones"
    worksheet.freeze_panes = 'A2'
    
    for row in worksheet.iter_rows(min_col=10 ,max_col=11, min_row=2):
        for cell in row:
            cell.style = date_style
    
    for index in range(13):
        worksheet[f'{get_column_letter(index + 1)}1'].fill = green
        
    automatic_width( worksheet = worksheet )
    
    worksheet['O1'].fill = red
    
    if BUSY_LOCATIONS:
        worksheet['O5'].fill = blue
    
    worksheet.column_dimensions['J'].width = 11
    worksheet.column_dimensions['K'].width = 11
    # worksheet.column_dimensions['L'].width = 10
    # worksheet.column_dimensions['M'].width = 20
    worksheet.column_dimensions['O'].width = 20
    worksheet.column_dimensions['P'].width = 20
    
    worksheet = workbook["Historia"]
    
    automatic_width(worksheet = worksheet)
    
    worksheet['A1'].fill = blue
    worksheet['B1'].fill = blue
    worksheet['D2'].fill = red
    
    worksheet.column_dimensions['C'].width = 5
    worksheet.column_dimensions['D'].width = 22
    
    workbook.save(EXCEL_FILE)


# Genera un resumen del mes y el grafico
def history():
    data = fetchAll()
    
    alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    workbook = load_workbook(EXCEL_FILE)
    workbook.create_sheet("Historia")
    
    worksheet = workbook["Historia"]
    
    average = 0
    index = 2
    for row in data:
        worksheet[f'A{index}'].value = row[1]
        worksheet[f'B{index}'].value = row[2]
        worksheet[f'A{index}'].alignment = alignment
        worksheet[f'B{index}'].alignment = alignment
        worksheet[f'A{index}'].border = border
        worksheet[f'B{index}'].border = border
        average += int(row[2])
        index +=1
        
    average = average/(index-2)
    average = math.ceil(average)
    
    worksheet['A1'].value = "Fecha"
    worksheet['B1'].value = "Almacenes ocupados"
    worksheet['D2'].value = "PROMEDIO"
    worksheet['D3'].value = average
    
    worksheet['A1'].alignment = alignment
    worksheet['B1'].alignment = alignment
    worksheet['D2'].alignment = alignment
    worksheet['D3'].alignment = alignment
    
    worksheet['A1'].border = border
    worksheet['B1'].border = border
    worksheet['D2'].border = border
    worksheet['D3'].border = border
        
    workbook.save(EXCEL_FILE)


# Genera un grafico
def createChart():
    locale.setlocale(locale.LC_TIME, 'es_ES')
    month = datetime.now().strftime('%B')
    
    df = pd.read_excel(EXCEL_FILE, sheet_name='Historia')
    df = df[['Fecha', 'Almacenes ocupados']]

    wb = load_workbook(EXCEL_FILE)
    ws = wb['Historia']
    
    df.dropna(inplace = True)

    plt.figure(figsize=(8, 4))
    plt.bar(df['Fecha'], df['Almacenes ocupados'], color='#B7DEE8')

    plt.title(f'Ocupacion del almacen({month})', fontsize=16, fontweight='bold')
    plt.xlabel('Fecha', fontsize=12)
    plt.ylabel('Almacenes ocupados', fontsize=12)

    plt.grid(axis='y', linestyle='--', alpha=0.7)

    plt.xticks(rotation=90)

    for i, valor in enumerate(df['Almacenes ocupados']):
        plt.text(i, valor + 1, str(valor), ha='center', fontsize=6, fontweight='bold')
        
    plt.tight_layout()
    
    grafica_path = os.path.join(FOLDER_PATH, "grafica.png")
    plt.savefig(grafica_path, dpi=300, bbox_inches='tight')

    img = Image(grafica_path)
    img.width = 750
    img.height = 450
    ws.add_image(img, 'F2')

    wb.save(EXCEL_FILE)
    



# Envia el correo
def email():
    subject = "INFORME DE UBICACIONES"

    emails = [
        "ferney.correa@prebel.com.co",
        "practicante.almacenamiento@prebel.com.co",
        # "jefedealmacen@grupomilagros.com", 
        # "jefeabastecimiento@grupomilagros.com", 
        # "tesoreria@grupomilagros.com", 
        # "costos@grupomilagros.com", 
        # "jefelogistica@grupomilagros.com", 
        # "coordinadora.compras@grupomilagros.com",
        ]


    copy_to = [
        # "fray.Henao@prebel.com.co",
        # "jose.otero@prebel.com.co",
        # "cesar.martinez@prebel.com.co",
        # "paula.correa@prebel.com.co",
        # "ferney.correa@prebel.com.co",
        # "practicante.almacenamiento@prebel.com.co",
    ]

    body = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Informe de ubicaciones</title>
        </head>
        <body>
            <h2>Cordial saludo</h2>
            <p>Adjunto el informe de ubicaciones</p>
            <br>
            <p>Saludos</p>
        </body>
        </html>
    """

    attachments = [
        EXCEL_FILE
    ]

    send_email(
        subject = subject,
        emails = "; ".join(emails),
        emails_cc = "; ".join(copy_to),
        files_Attachment = attachments,
        html_contend = body
    )
    

# Inicia el flujo
def main():
    
    try:
        today = date.today()
        
        create_folder("Informe de ubicaciones")
        close_excel(EXCEL_FILE)
            
        if AFK:
            sap_instance = SapGui(
                conection = "PRD [PRODUCTIVO]",
                user = USER,
                password = PASSWORD,
            )

        else:
            sap_instance = SapGui(
                conection = "PRD [PRODUCTIVO]",
                active = True
            )
            
        main_df = LX02( sap_instance = sap_instance )
        df_zpp56 = ZPP56(sap_instance = sap_instance, main_df = main_df)
        
        if AFK:
            sap_instance.close_sap()
        
        data = processData( main_df = main_df, df_zpp56 = df_zpp56)
        
        mainDataToExcel( df = data, date = today )
        SummaryToExcel( df = data, date = today  )
        
        history()
        add_no_facturar_sheet()
        addStyles()
        createChart()
        
        if SEND_EMAILS:
            email()
           
        if AFK: 
            report_status(
                script= "informe_ubicacion",
                status = True,
            )
    
    except Exception as e:
        
        if AFK:
            report_status(
                script = "informe_ubicaciones",
                status = False,
                message = f"ERROR {e}"
            )
            
        print(e)


main()

#<────────────────────────────────────────────────────────────────────────────────>

#Script para insertar datos a sqlite (informe ubicaciones)
# connection = sqlite3.connect(DATABASE_PATH)
# cursor = connection.cursor()

# cursor.execute('''
#           INSERT INTO ocupacion (fecha, ubicaciones_ocupadas) 
#            VALUES ('2025-03-04', 65)
#     ''')
# connection.commit()   # Guarda los cambios realizados en la base de datos.   
# connection.close()    # Cierra la conexion a la base de datos.

#<────────────────────────────────────────────────────────────────────────────────>

#Script para borrar datos de sqlite (informe ubicaciones)
# connection = sqlite3.connect(DATABASE_PATH)
# cursor = connection.cursor()
# cursor.execute('''
#     DELETE 
#     FROM ocupacion
#     WHERE id = 355
#     ''')
# connection.commit()   # --> Guarda los cambios realizados en la base de datos.
# connection.close()    # --> Cierra la conexion a la base de datos.        

#<────────────────────────────────────────────────────────────────────────────────>    



from sap_gui_library import SapGui, Transaction, DataProcess
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from colorama import init, Fore, Style
from Generales import create_folder, close_excel, automatic_width
from datetime import datetime
import pandas as pd
import re,os

# Constantes
PATH = "C:\TEMP\Traslados_destelle"

# Validar el formato de las fechas
def validate_date(date:str):
    
    # Patron para validar que la fecha tenga formato dd.mm.aaaa
    pattern = r"^\d{2}\.\d{2}\.\d{4}$"
    
    # re.match valida que el formato corresponda al str recibido
    if not re.match(pattern, date):
        return False
    
    # Si la fecha esta en el calendario, podra convertirse a datetime
    try:
        datetime.strptime(date, '%d.%m.%Y')
        return True
    
    except ValueError:
        return False
    

# Ejecuta la transaccion ZWM85 para obtener los datos
def ZWM85( sap_instance:SapGui, start_date:str, end_date:str ) -> pd.DataFrame:
    
    # inicializar la transaccion
    transaction = Transaction(
        session = sap_instance.session,
        code = "ZWM85"
    )
    
    transaction.start_transaction()
    
    # Parametros de la transaccion
    transaction.session.findById("wnd[0]/usr/ctxtSO_QDATU-LOW").text = start_date
    transaction.session.findById("wnd[0]/usr/ctxtSO_QDATU-HIGH").text = end_date
    transaction.session.findById("wnd[0]/usr/ctxtSO_QZEIT-LOW").text = "00:00:00"
    transaction.session.findById("wnd[0]/usr/ctxtSO_QZEIT-HIGH").text = "23:59:00"
    transaction.session.findById("wnd[0]/usr/ctxtSO_LGNUM-LOW").text = "pro"
    transaction.session.findById("wnd[0]/usr/ctxtSO_VLTYP-LOW").text = "app"
    transaction.session.findById("wnd[0]/usr/ctxtSO_NLTYP-LOW").text = "asp"
    
    transaction.run_transaction( bttn = 'Execute')
    
    # Pasa la informacion a lista plana
    transaction.session.findById("wnd[0]/usr/cntlALV_CONTAINER/shellcont/shell/shellcont[1]/shell[0]").pressButton("DETAIL")
    transaction.session.findById("wnd[1]/tbar[0]/btn[46]").press()
    
    # Exporta y descarga en formato csv
    transaction.export_in_toolbar(45)
    
    transaction.select_export_and_download(
        address = PATH,
        name = "Traslados_destelle"
    )
    
    # Procesa el archivo csv
    dataprocess = DataProcess(
        address_file = os.path.join(PATH, "Traslados_destelle.csv"),
        skiprows = 1
    )
    
    # Edita¿ el archivo csv
    dataprocess.edit_file(
        skip_first_row = True,
        strip_columns = True,
        columns_from_str_to_float = ['Unid.', 'Número OT']
    )
    
    # Convierte el archivo csv a un DataFrame
    df = dataprocess.get_df()
    
    # Guarda el DataFrame en un archivo Excel
    df.to_excel(os.path.join(PATH, "Traslados_destelle.xlsx"), index = False)
    
    return df


# Procesa los datos obtenidos y muestra el resultado
def processData( df:pd.DataFrame ):
    print(
        f"\n<-------------------- {Fore.BLUE}PROCESAMIENTO DE DATOS{Style.RESET_ALL} -------------------->\n"
        f"Registros totales: {Fore.GREEN}"
        f"{len(df)}{Style.RESET_ALL}\n"
        
        # Elimina los registros duplicados en la columna 'Numero OT'
        f"Registros filtrados: {Fore.GREEN}"
        f"{len(df.drop_duplicates(subset = ['Número OT']))}{Style.RESET_ALL} \n"
    )


# Estiliza el archivo de Excel
def generate_excel():
    blue = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    workbook = load_workbook(os.path.join(PATH, "Traslados_destelle.xlsx"))
    worksheet = workbook['Sheet1']
    
    # Ajusta el ancho de las columnas
    automatic_width(worksheet)
    
    # Cambia el color de los encabezados
    for index in range(1, 13):
        worksheet[f'{get_column_letter(index)}1'].fill = blue
        
    workbook.save(os.path.join(PATH, "Traslados_destelle.xlsx"))
    

# Flujo principal
def main():
    # Inicializa el modulo colorama
    init()
    
    # Intenta cerrar el archivo Excel, en caso de que este abierto
    while True:
        try:
            close_excel( file_path = os.path.join(PATH, "Traslados_destelle.xlsx"))
            break
        
        except:
            print(
                f"{Fore.YELLOW}ADVERTENCIA: {Style.RESET_ALL}Es necesario cerrar el libro de Excel ->"
                f"{Fore.CYAN}Traslados_destelle.xlsx{Style.RESET_ALL}"
            )
            
            input("Presione Enter para continuar...\n")
            os.system("cls")
            
    # Crea la carpeta donde se guardara el archivo Excel
    create_folder("Traslados_destelle")
    
    # Inicializa la instancia de SAP
    while True:
        try:
            sap_instance = SapGui( active = True )
            break
        
        except:
            print(
                f"{Fore.YELLOW}ADVERTENCIA: {Style.RESET_ALL}Es necesario iniciar sesion en SAP"
            )
            
            input("Presione Enter para continuar...\n")
            os.system("cls")
    
    # Solicita el rango de fechas
    print(
        f"<-------------------- {Fore.BLUE}RANGO DE FECHAS{Style.RESET_ALL} -------------------->\n"
        f"ingrese el rango de fechas en formato -> {Fore.GREEN}"
        f"dd.mm.aaaa {Style.RESET_ALL}\n"
    )
    
    start_date = input(">> Fecha de inicio: ")
    end_date = input(">> Fecha de finalizacion: ")
    
    # Valida el formato de las fechas
    if not validate_date(start_date) or not validate_date(end_date):
        print(
            f"{Fore.RED}El formato de las fechas no es correcto\n"
            f"Por favor, intente de nuevo{Style.RESET_ALL}"
        )
        
        input("Presione Enter para continuar...")
        os.system("cls")
        main()
        
        return
    
    # Ejecuta la transaccion ZWM85
    ZWM85_df = ZWM85(
        start_date = start_date,
        end_date = end_date,
        sap_instance = sap_instance
    )
    
    # Procesa los datos obtenidos
    processData( df = ZWM85_df )
    generate_excel()
    
    input( ">> Presione Enter para finalizar...")
    

if __name__ == "__main__":    
    main()